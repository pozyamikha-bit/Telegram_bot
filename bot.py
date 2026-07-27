"""
Бот регистрации участников акции + приём чеков/УПД + админ-панель.

Версия под aiogram 3.x (актуальная ветка библиотеки; именно её ожидает
и платформа Amvera при сборке — см. README.md).

Сценарий для обычного пользователя:
  /start -> ФИО -> Магазин -> Телефон (кнопкой) -> запись в Google Таблицу
  -> главное меню -> "Отправить чек" -> фото -> запись в Google Таблицу

Админ-панель (/admin, доступно владельцам из ADMIN_IDS и модераторам,
добавленным через саму панель):
  - "История" — календарь, по выбранной дате список всех чеков за неё
    с инфо об отправителе и фото.
  - "Модерация" — список чеков со статусом "на модерации"; для каждого
    можно принять (после чего запрашивается номер купона и он
    отправляется пользователю) или отклонить (пользователь уведомляется).
  - "Модераторы" (только для владельцев из ADMIN_IDS) — добавление и
    удаление модераторов.
  - "Отчёт" — Excel-файл со всеми чеками и регистрациями.
"""

import asyncio
import calendar
import io
import logging
import os
import re
import datetime
from datetime import date

import openpyxl
from openpyxl.utils import get_column_letter

from aiogram import Bot, Dispatcher, F
from aiogram.filters import BaseFilter, Command, CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    Message,
)
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder

import config
import google_sheets

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
logger = logging.getLogger(__name__)

bot = Bot(token=config.BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())

os.makedirs(config.RECEIPTS_DIR, exist_ok=True)


class Form(StatesGroup):
    name = State()
    shop = State()
    inn = State()
    phone = State()
    waiting_photo = State()
    waiting_upd_number = State()
    waiting_receipt_date = State()


class AdminForm(StatesGroup):
    accept_coupon_choice = State()
    waiting_coupon = State()
    waiting_moderator_id = State()
    waiting_text_value = State()
    waiting_reject_reason = State()
    waiting_promo_file = State()
    waiting_sales_file = State()
    waiting_coupons_file = State()
    picking_period = State()
    auto_mod_confirm = State()
    waiting_inn_fix = State()
    coupons_confirm = State()


# ---------- Проверка прав доступа ----------

def _is_owner(user_id: int) -> bool:
    return user_id in config.ADMIN_IDS


async def _is_moderator(user_id: int) -> bool:
    if _is_owner(user_id):
        return True
    try:
        return user_id in google_sheets.get_moderator_ids()
    except Exception:
        logger.exception("Не удалось проверить список модераторов")
        return False


class IsOwner(BaseFilter):
    async def __call__(self, event) -> bool:
        return _is_owner(event.from_user.id)


class IsModerator(BaseFilter):
    async def __call__(self, event) -> bool:
        return await _is_moderator(event.from_user.id)


# ---------- Клавиатуры ----------

BTN_SEND_RECEIPT = "📥 Отправить чек / УПД"
BTN_MY_RECEIPTS = "📄 Загруженные чеки"
BTN_DELETE_RECEIPTS = "🗑 Удалить чеки"
BTN_RULES = "❓ Правила"

ADMIN_BTN_HISTORY = "🗓 История"
ADMIN_BTN_MODERATION = "🛠 Модерация"
ADMIN_BTN_MODERATORS = "👥 Модераторы"
ADMIN_BTN_REPORT = "📊 Отчёты"
ADMIN_BTN_TEXTS = "✏️ Тексты"
ADMIN_BTN_IMPORT = "📥 Импорт"
ADMIN_BTN_TEMPLATES = "🗂 Шаблоны"
ADMIN_BTN_INSTRUCTIONS = "📖 Инструкция"
ADMIN_BTN_EXIT = "⬅️ Выйти из панели"

_main_menu_builder = ReplyKeyboardBuilder()
_main_menu_builder.button(text=BTN_SEND_RECEIPT)
_main_menu_builder.button(text=BTN_MY_RECEIPTS)
_main_menu_builder.button(text=BTN_DELETE_RECEIPTS)
_main_menu_builder.button(text=BTN_RULES)
_main_menu_builder.adjust(1)
main_menu = _main_menu_builder.as_markup(resize_keyboard=True)

_moderator_menu_builder = ReplyKeyboardBuilder()
_moderator_menu_builder.button(text=ADMIN_BTN_HISTORY)
_moderator_menu_builder.button(text=ADMIN_BTN_MODERATION)
_moderator_menu_builder.button(text=ADMIN_BTN_REPORT)
_moderator_menu_builder.button(text=ADMIN_BTN_INSTRUCTIONS)
_moderator_menu_builder.button(text=ADMIN_BTN_EXIT)
_moderator_menu_builder.adjust(2, 2, 1)
moderator_menu = _moderator_menu_builder.as_markup(resize_keyboard=True)

_owner_menu_builder = ReplyKeyboardBuilder()
_owner_menu_builder.button(text=ADMIN_BTN_HISTORY)
_owner_menu_builder.button(text=ADMIN_BTN_MODERATION)
_owner_menu_builder.button(text=ADMIN_BTN_REPORT)
_owner_menu_builder.button(text=ADMIN_BTN_MODERATORS)
_owner_menu_builder.button(text=ADMIN_BTN_TEXTS)
_owner_menu_builder.button(text=ADMIN_BTN_IMPORT)
_owner_menu_builder.button(text=ADMIN_BTN_TEMPLATES)
_owner_menu_builder.button(text=ADMIN_BTN_INSTRUCTIONS)
_owner_menu_builder.button(text=ADMIN_BTN_EXIT)
_owner_menu_builder.adjust(2, 2, 2, 2, 1)
owner_menu = _owner_menu_builder.as_markup(resize_keyboard=True)


def _menu_for(user_id: int):
    return owner_menu if _is_owner(user_id) else moderator_menu


# ---------- Регистрация ----------

@dp.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()

    try:
        existing = google_sheets.get_registration_by_telegram_id(message.from_user.id)
    except Exception:
        logger.exception("Не удалось проверить регистрацию пользователя при /start")
        existing = None

    if existing:
        full_name = (existing.get("full_name") or "").strip()
        greeting = f"С возвращением, {full_name}!" if full_name else "С возвращением!"
        await message.answer(
            f"{greeting} Вы уже зарегистрированы, повторно проходить "
            "регистрацию не нужно.\n\n"
            "Можно отправить новый чек / УПД или посмотреть чеки, "
            "загруженные ранее.",
            reply_markup=main_menu,
        )
        return

    await state.set_state(Form.name)
    await message.answer(
        "Здравствуйте! Давайте зарегистрируем вас в программе.\n\n"
        "Для начала напишите, пожалуйста, ваше ФИО."
    )


# ---------- Вход в админ-панель (регистрируется рано, чтобы перебивать FSM) ----------

@dp.message(Command("admin"), IsModerator())
async def cmd_admin(message: Message, state: FSMContext):
    await state.clear()
    if _is_owner(message.from_user.id):
        await message.answer("Админ-панель.", reply_markup=owner_menu)
    else:
        await message.answer("Панель модератора.", reply_markup=moderator_menu)


@dp.message(F.text == ADMIN_BTN_EXIT, IsModerator())
async def admin_exit(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Вышли из панели.", reply_markup=main_menu)


@dp.message(F.text == ADMIN_BTN_INSTRUCTIONS, IsModerator())
async def admin_instructions(message: Message):
    # Два отдельных сообщения: инструкция по самой админ-панели и
    # инструкция по тому, как пользователь отправляет чек. Тексты
    # редактируются владельцем через "✏️ Тексты" (ключи instructions_admin
    # и instructions_receipt), поэтому их можно менять без изменения кода.
    await message.answer(google_sheets.get_text("instructions_admin"))
    await message.answer(google_sheets.get_text("instructions_receipt"))


# ---------- Регистрация (продолжение сценария) ----------

@dp.message(StateFilter(Form.name))
async def process_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text)
    await state.set_state(Form.shop)
    await message.answer("Отлично! Теперь укажите название вашего магазина.")


@dp.message(StateFilter(Form.shop))
async def process_shop(message: Message, state: FSMContext):
    await state.update_data(shop=message.text)
    await state.set_state(Form.inn)
    await message.answer(
        "Теперь укажите ИНН магазина (10 цифр — для организации, "
        "или 12 цифр — для ИП)."
    )


_INN_RE = re.compile(r"^\d{10}$|^\d{12}$")


@dp.message(StateFilter(Form.inn), F.text)
async def process_inn(message: Message, state: FSMContext):
    inn = message.text.strip()

    if not _INN_RE.match(inn):
        await message.answer(
            "ИНН должен состоять только из цифр — 10 цифр для организации "
            "или 12 цифр для ИП. Введите ИНН ещё раз."
        )
        return

    await state.update_data(inn=inn)

    kb_builder = ReplyKeyboardBuilder()
    kb_builder.button(text="Поделиться контактом", request_contact=True)
    keyboard = kb_builder.as_markup(resize_keyboard=True, one_time_keyboard=True)

    await state.set_state(Form.phone)
    await message.answer(
        "И последнее: поделись номером телефона для связи.",
        reply_markup=keyboard,
    )


@dp.message(StateFilter(Form.inn))
async def wrong_content_for_inn(message: Message):
    await message.answer("Пожалуйста, введите ИНН магазина текстом (только цифры).")


@dp.message(StateFilter(Form.phone), F.contact)
async def process_phone(message: Message, state: FSMContext):
    data = await state.get_data()
    full_name = data.get("name", "")
    shop = data.get("shop", "")
    inn = data.get("inn", "")
    phone = message.contact.phone_number

    try:
        google_sheets.append_registration(
            full_name=full_name,
            shop=shop,
            inn=inn,
            phone=phone,
            telegram_id=message.from_user.id,
            username=message.from_user.username,
        )
    except Exception:
        logger.exception("Не удалось записать регистрацию в Google Таблицу")
        await message.answer(
            "Регистрация прошла, но возникла ошибка при записи в таблицу. "
            "Сообщите об этом администратору."
        )

    await state.clear()
    await message.answer(
        google_sheets.get_text("registration_success"),
        reply_markup=main_menu,
    )


# ---------- Правила ----------

@dp.message(F.text == BTN_RULES)
async def show_rules(message: Message):
    await message.answer(google_sheets.get_text("rules"))


# ---------- Приём чеков ----------

RECEIPT_CANCEL_CALLBACK = "receipt_cancel"


def _cancel_kb(text: str = "❌ Отмена"):
    kb = InlineKeyboardBuilder()
    kb.row(InlineKeyboardButton(text=text, callback_data=RECEIPT_CANCEL_CALLBACK))
    return kb.as_markup()


def _calendar_with_cancel(year: int, month: int, prefix: str = "rcal"):
    markup = _build_calendar(year, month, prefix=prefix)
    markup.inline_keyboard.append(
        [InlineKeyboardButton(text="❌ Отмена", callback_data=RECEIPT_CANCEL_CALLBACK)]
    )
    return markup


@dp.message(F.text == BTN_SEND_RECEIPT)
async def send_receipt_start(message: Message, state: FSMContext):
    await state.set_state(Form.waiting_photo)
    await message.answer(google_sheets.get_text("ask_photo"), reply_markup=_cancel_kb())


@dp.message(StateFilter(Form.waiting_photo), F.photo)
async def process_photo(message: Message, state: FSMContext):
    photo = message.photo[-1]  # фото в максимальном качестве
    file_id = photo.file_id

    filename = f"{message.from_user.id}_{message.date.strftime('%Y%m%d_%H%M%S')}.jpg"
    filepath = os.path.join(config.RECEIPTS_DIR, filename)

    try:
        await bot.download(file_id, destination=filepath)
    except Exception:
        logger.exception("Не удалось скачать фото чека")
        filepath = "(не скачалось)"

    # Фото ещё не сохраняем в таблицу — сначала нужно получить от
    # пользователя номер УПД и дату чека вручную, чтобы записать всё
    # одной строкой.
    await state.update_data(photo_file_id=file_id, photo_file_name=filepath)
    await state.set_state(Form.waiting_upd_number)
    await message.answer(google_sheets.get_text("ask_upd_number"), reply_markup=_cancel_kb())


@dp.message(StateFilter(Form.waiting_photo))
async def wrong_content_for_photo(message: Message):
    await message.answer(
        "Пожалуйста, пришлите именно фото (как изображение, не файлом).",
        reply_markup=_cancel_kb(),
    )


@dp.message(StateFilter(Form.waiting_upd_number), F.text)
async def process_upd_number(message: Message, state: FSMContext):
    upd_number = message.text.strip()

    try:
        duplicate = google_sheets.find_active_receipt_by_upd(upd_number)
    except Exception:
        logger.exception("Не удалось проверить № УПД на повтор")
        duplicate = None

    if duplicate:
        await message.answer(
            "Этот номер УПД уже был подан ранее (один чек — один купон "
            "Ozon, повторно использовать тот же № УПД нельзя). Проверьте "
            "номер и введите его ещё раз, либо нажмите «Отмена».",
            reply_markup=_cancel_kb(),
        )
        return

    await state.update_data(upd_number=upd_number)
    await state.set_state(Form.waiting_receipt_date)
    today = google_sheets.moscow_today()
    await message.answer(
        google_sheets.get_text("ask_receipt_date"),
        reply_markup=_calendar_with_cancel(today.year, today.month, prefix="rcal"),
    )


@dp.message(StateFilter(Form.waiting_upd_number))
async def wrong_content_for_upd_number(message: Message):
    await message.answer(
        "Пожалуйста, введите номер УПД текстом (одним сообщением).",
        reply_markup=_cancel_kb(),
    )


@dp.callback_query(F.data == RECEIPT_CANCEL_CALLBACK)
async def receipt_cancel(call: CallbackQuery, state: FSMContext):
    current_state = await state.get_state()
    active_states = {
        Form.waiting_photo.state,
        Form.waiting_upd_number.state,
        Form.waiting_receipt_date.state,
    }
    if current_state not in active_states:
        await call.answer("Загрузка чека уже завершена или отменена.", show_alert=True)
        return

    await state.clear()
    try:
        await call.message.edit_reply_markup(reply_markup=None)
    except Exception:
        pass
    await call.answer("Отменено.")
    await call.message.answer(
        "Загрузка чека отменена. Можно отправить чек заново или "
        "посмотреть уже загруженные чеки.",
        reply_markup=main_menu,
    )


@dp.callback_query(F.data == "rcal_ignore")
async def receipt_calendar_ignore(call: CallbackQuery):
    await call.answer()


@dp.callback_query(F.data.startswith("rcal_nav:"), StateFilter(Form.waiting_receipt_date))
async def receipt_calendar_nav(call: CallbackQuery):
    _, ym = call.data.split(":", 1)
    year, month = map(int, ym.split("-"))
    await call.message.edit_reply_markup(reply_markup=_calendar_with_cancel(year, month, prefix="rcal"))
    await call.answer()


@dp.callback_query(F.data.startswith("rcal_day:"), StateFilter(Form.waiting_receipt_date))
async def receipt_calendar_pick_day(call: CallbackQuery, state: FSMContext):
    _, date_str = call.data.split(":", 1)
    year_s, month_s, day_s = date_str.split("-")
    receipt_date = f"{day_s}.{month_s}.{year_s}"

    data = await state.get_data()
    upd_number = data.get("upd_number", "")

    # Повторная проверка прямо перед записью — на случай, если пока
    # пользователь выбирал дату, кто-то другой успел подать тот же № УПД
    # (проверка на шаге ввода номера УПД саму по себе не гарантирует
    # отсутствие "гонки" между двумя параллельными заявками).
    try:
        duplicate = google_sheets.find_active_receipt_by_upd(upd_number)
    except Exception:
        logger.exception("Не удалось повторно проверить № УПД перед записью")
        duplicate = None

    if duplicate:
        await state.clear()
        await call.message.edit_reply_markup(reply_markup=None)
        await call.message.answer(
            "Этот номер УПД уже был подан (кто-то успел раньше). Загрузка "
            "чека отменена — если считаете это ошибкой, обратитесь к "
            "администратору.",
            reply_markup=main_menu,
        )
        await call.answer()
        return

    try:
        google_sheets.append_receipt(
            telegram_id=call.from_user.id,
            username=call.from_user.username,
            file_id=data.get("photo_file_id", ""),
            file_name=data.get("photo_file_name", ""),
            upd_number=upd_number,
            receipt_date=receipt_date,
        )
    except Exception:
        logger.exception("Не удалось записать чек в Google Таблицу")

    await state.clear()
    await call.message.edit_reply_markup(reply_markup=None)
    await call.message.answer(
        google_sheets.get_text("receipt_received"),
        reply_markup=main_menu,
    )
    await call.answer(f"Дата чека: {receipt_date}")


@dp.callback_query(F.data.startswith("rcal_day:"))
async def receipt_calendar_pick_day_stale(call: CallbackQuery):
    await call.answer(
        "Эта кнопка уже не активна. Начните заново через «📥 Отправить чек / УПД».",
        show_alert=True,
    )


@dp.message(StateFilter(Form.waiting_receipt_date))
async def wrong_content_for_receipt_date(message: Message):
    await message.answer("Пожалуйста, выберите дату чека кнопками календаря в сообщении выше.")


# ---------- Загруженные чеки (для самого пользователя) ----------

@dp.message(F.text == BTN_MY_RECEIPTS)
async def my_receipts(message: Message):
    await message.answer("Формирую отчёт по вашим чекам, секунду...")
    try:
        receipts = google_sheets.get_receipts_by_telegram_id(message.from_user.id)
    except Exception:
        logger.exception("Не удалось получить чеки пользователя из Google Таблицы")
        await message.answer("Не получилось сформировать отчёт, попробуйте позже.")
        return

    if not receipts:
        await message.answer("Вы ещё не загружали ни одного чека.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Мои чеки"
    headers = ["Дата загрузки", "№ УПД", "Дата чека", "Статус", "Купон", "Комментарий"]
    ws.append(headers)
    for r in receipts:
        ws.append([
            r["date"],
            r["upd_number"],
            r["receipt_date"],
            r["status"],
            r["coupon"],
            r["comment"],
        ])
    _autosize(ws, headers)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"moi_cheki_{message.from_user.id}.xlsx"
    await message.answer_document(BufferedInputFile(buf.read(), filename=filename))


# ---------- Удаление своих чеков (для самого пользователя) ----------

async def _send_own_receipts_list(message: Message, user_id: int):
    try:
        receipts = google_sheets.get_receipts_by_telegram_id(user_id, include_deleted=False)
    except Exception:
        logger.exception("Не удалось получить чеки пользователя для удаления")
        await message.answer("Не получилось получить список чеков, попробуйте позже.")
        return

    # Удалять самому можно только чеки, которые ещё НЕ прошли модерацию
    # (статус "на модерации") — принятые/отклонённые чеки менять задним
    # числом пользователю нельзя (по ним уже могли отправить купон или
    # отказ, это должно остаться в таблице).
    receipts = [r for r in receipts if r["status"] == google_sheets.STATUS_PENDING]

    if not receipts:
        await message.answer(
            "У вас нет чеков, которые можно удалить — удалить можно "
            "только чеки со статусом «на модерации» (уже "
            "промодерированные принятые/отклонённые чеки удалить нельзя)."
        )
        return

    kb = InlineKeyboardBuilder()
    for r in receipts:
        parts = []
        if r.get("upd_number"):
            parts.append(f"УПД {r['upd_number']}")
        else:
            parts.append("без № УПД")
        if r.get("receipt_date"):
            parts.append(r["receipt_date"])
        label = _truncate(" · ".join(parts), 60)
        kb.row(InlineKeyboardButton(text=label, callback_data=f"my_del_view:{r['row']}"))

    await message.answer("Выберите чек, который нужно удалить:", reply_markup=kb.as_markup())


@dp.message(F.text == BTN_DELETE_RECEIPTS)
async def my_receipts_delete_menu(message: Message):
    await _send_own_receipts_list(message, message.from_user.id)


async def _send_own_receipt_card(message: Message, receipt: dict):
    lines = [
        f"№ УПД: {receipt['upd_number'] or '—'}",
        f"Дата чека: {receipt['receipt_date'] or '—'}",
        f"Дата загрузки: {receipt['date']}",
        f"Статус: {receipt['status']}",
    ]
    if receipt.get("coupon"):
        lines.append(f"Купон: {receipt['coupon']}")
    if receipt.get("comment"):
        lines.append(f"Комментарий: {receipt['comment']}")
    caption = "\n".join(lines)

    kb = InlineKeyboardBuilder()
    kb.row(InlineKeyboardButton(text="🗑 Удалить чек", callback_data=f"my_del_confirm:{receipt['row']}"))
    kb.row(InlineKeyboardButton(text="⬅️ Назад к списку", callback_data="my_del_back"))
    reply_markup = kb.as_markup()

    if receipt["file_id"]:
        try:
            await message.answer_photo(receipt["file_id"], caption=caption, reply_markup=reply_markup)
            return
        except Exception:
            logger.exception("Не удалось отправить фото чека пользователю (удаление чеков)")

    await message.answer(caption, reply_markup=reply_markup)


def _owns_receipt(receipt: dict, user_id: int) -> bool:
    return bool(receipt) and receipt["telegram_id"] == str(user_id)


def _user_can_delete_receipt(receipt: dict) -> bool:
    """Пользователь может удалить только свой ещё не промодерированный
    чек (статус "на модерации") — принятые/отклонённые уже нельзя."""
    return bool(receipt) and not receipt["deleted"] and receipt["status"] == google_sheets.STATUS_PENDING


@dp.callback_query(F.data.startswith("my_del_view:"))
async def my_receipts_delete_view(call: CallbackQuery):
    row = int(call.data.split(":", 1)[1])
    receipt = google_sheets.get_receipt_by_row(row)
    if not _owns_receipt(receipt, call.from_user.id):
        await call.answer("Чек не найден.", show_alert=True)
        return
    if receipt["deleted"]:
        await call.answer("Этот чек уже удалён.", show_alert=True)
        return
    if not _user_can_delete_receipt(receipt):
        await call.answer(
            "Этот чек уже прошёл модерацию — удалить его самостоятельно нельзя.",
            show_alert=True,
        )
        return

    await _send_own_receipt_card(call.message, receipt)
    await call.answer()


@dp.callback_query(F.data == "my_del_back")
async def my_receipts_delete_back(call: CallbackQuery):
    await call.answer()
    await _send_own_receipts_list(call.message, call.from_user.id)


@dp.callback_query(F.data.startswith("my_del_confirm:"))
async def my_receipts_delete_confirm(call: CallbackQuery):
    row = int(call.data.split(":", 1)[1])
    receipt = google_sheets.get_receipt_by_row(row)
    if not _owns_receipt(receipt, call.from_user.id):
        await call.answer("Чек не найден.", show_alert=True)
        return
    if not _user_can_delete_receipt(receipt):
        await call.answer(
            "Этот чек уже нельзя удалить (уже удалён или прошёл модерацию).",
            show_alert=True,
        )
        return

    kb = InlineKeyboardBuilder()
    kb.row(
        InlineKeyboardButton(text="✅ Да, удалить", callback_data=f"my_del_yes:{row}"),
        InlineKeyboardButton(text="❌ Отмена", callback_data=f"my_del_view:{row}"),
    )
    await call.message.answer(
        "Точно удалить этот чек? Действие необратимо — чек больше не "
        "будет учитываться при сверке с отчётом по продажам.",
        reply_markup=kb.as_markup(),
    )
    await call.answer()


@dp.callback_query(F.data.startswith("my_del_yes:"))
async def my_receipts_delete_do(call: CallbackQuery):
    row = int(call.data.split(":", 1)[1])
    receipt = google_sheets.get_receipt_by_row(row)
    if not _owns_receipt(receipt, call.from_user.id):
        await call.answer("Чек не найден.", show_alert=True)
        return
    if not _user_can_delete_receipt(receipt):
        await call.answer(
            "Этот чек уже нельзя удалить (уже удалён или прошёл модерацию).",
            show_alert=True,
        )
        return

    try:
        google_sheets.mark_receipt_deleted(row)
    except Exception:
        logger.exception("Не удалось удалить чек пользователя")
        await call.answer("Не получилось удалить, попробуйте позже.", show_alert=True)
        return

    await call.answer("Чек удалён.")
    await call.message.answer("Готово, чек удалён и больше не участвует в сверке.")
    await _send_own_receipts_list(call.message, call.from_user.id)


# ---------- Общая карточка чека (используется и в Истории, и в Модерации) ----------

async def _send_receipt_card(
    message: Message,
    receipt: dict,
    with_moderation_buttons: bool = False,
    with_history_buttons: bool = False,
    viewer_id: int = None,
):
    reg = google_sheets.get_registration_by_telegram_id(receipt["telegram_id"])

    lines = []
    if receipt.get("deleted"):
        lines.append("🗑 Помечен как удалённый")
    if reg:
        lines.append(f"ФИО: {reg['full_name']}")
        lines.append(f"Магазин: {reg['shop']}")
        if reg.get("inn"):
            lines.append(f"ИНН магазина: {reg['inn']}")
        lines.append(f"Телефон: {reg['phone']}")
    else:
        lines.append("Регистрация не найдена (данные могли не сохраниться).")

    lines.append(f"Telegram ID: {receipt['telegram_id']}")
    if receipt["username"]:
        lines.append(f"Username: @{receipt['username']}")
    lines.append(f"Дата загрузки: {receipt['date']}")
    if receipt.get("upd_number"):
        lines.append(f"№ УПД (введён пользователем): {receipt['upd_number']}")
    if receipt.get("receipt_date"):
        lines.append(f"Дата чека (введена пользователем): {receipt['receipt_date']}")
    lines.append(f"Статус: {receipt['status']}")
    if receipt.get("coupon"):
        lines.append(f"Купон: {receipt['coupon']}")
    if receipt.get("comment"):
        lines.append(f"Комментарий: {receipt['comment']}")

    caption = "\n".join(lines)

    kb = InlineKeyboardBuilder()
    has_buttons = False

    # Кнопки модерации (Принять / Быстрое отклонение / Отклонить с
    # комментарием) показываем и в самой Модерации, и в Истории — из
    # Истории администратор тоже может сразу принять или отклонить чек,
    # не переходя отдельно в "🛠 Модерация".
    if with_moderation_buttons or with_history_buttons:
        kb.row(InlineKeyboardButton(text="✅ Принять", callback_data=f"mod_accept:{receipt['row']}"))
        kb.row(InlineKeyboardButton(
            text="⚡ Быстрое отклонение",
            callback_data=f"mod_reject_photo:{receipt['row']}",
        ))
        kb.row(InlineKeyboardButton(
            text="💬 Отклонить с комментарием",
            callback_data=f"mod_reject_custom:{receipt['row']}",
        ))
        has_buttons = True

    if with_history_buttons:
        kb.row(InlineKeyboardButton(text="🗑 Удалить чек", callback_data=f"hist_delete:{receipt['row']}"))
        date_str = receipt["date"][:10] if receipt.get("date") else ""
        if date_str:
            kb.row(InlineKeyboardButton(text="⬅️ Назад к списку", callback_data=f"cal_day:{date_str}"))
        has_buttons = True

    if reg and viewer_id is not None and _is_owner(viewer_id):
        kb.row(InlineKeyboardButton(
            text="✏️ Исправить ИНН",
            callback_data=f"fix_inn:{receipt['telegram_id']}",
        ))
        has_buttons = True

    reply_markup = kb.as_markup() if has_buttons else None

    if receipt["file_id"]:
        try:
            await message.answer_photo(receipt["file_id"], caption=caption, reply_markup=reply_markup)
            return
        except Exception:
            logger.exception("Не удалось отправить фото чека администратору")

    await message.answer(caption, reply_markup=reply_markup)


# ---------- Исправление ИНН пользователя (владелец, прямо с карточки чека) ----------

@dp.callback_query(F.data.startswith("fix_inn:"), IsOwner())
async def admin_fix_inn_start(call: CallbackQuery, state: FSMContext):
    target_telegram_id = call.data.split(":", 1)[1]
    await state.set_state(AdminForm.waiting_inn_fix)
    await state.update_data(fix_inn_telegram_id=target_telegram_id)
    await call.message.answer(
        f"Введите новый ИНН магазина для пользователя с Telegram ID "
        f"{target_telegram_id} (10 цифр для организации или 12 для ИП)."
    )
    await call.answer()


@dp.message(StateFilter(AdminForm.waiting_inn_fix), F.text, IsOwner())
async def admin_fix_inn_finish(message: Message, state: FSMContext):
    new_inn = message.text.strip()

    if not _INN_RE.match(new_inn):
        await message.answer(
            "ИНН должен состоять только из цифр — 10 цифр для организации "
            "или 12 цифр для ИП. Введите ещё раз."
        )
        return

    data = await state.get_data()
    target_telegram_id = data.get("fix_inn_telegram_id")
    await state.clear()

    try:
        updated = google_sheets.update_registration_inn(target_telegram_id, new_inn)
    except Exception:
        logger.exception("Не удалось обновить ИНН пользователя %s", target_telegram_id)
        await message.answer(
            "Не получилось обновить ИНН, попробуйте позже.",
            reply_markup=_menu_for(message.from_user.id),
        )
        return

    if not updated:
        await message.answer(
            "Не нашёл регистрацию с таким Telegram ID — возможно, она была удалена.",
            reply_markup=_menu_for(message.from_user.id),
        )
        return

    await message.answer(
        f"Готово, ИНН обновлён на «{new_inn}».",
        reply_markup=_menu_for(message.from_user.id),
    )


@dp.message(StateFilter(AdminForm.waiting_inn_fix), IsOwner())
async def wrong_content_for_inn_fix(message: Message):
    await message.answer("Пожалуйста, введите новый ИНН текстом (только цифры).")


def _truncate(text: str, max_len: int) -> str:
    text = text or ""
    if len(text) > max_len:
        return text[: max_len - 1] + "…"
    return text


def _format_receipt_label(r: dict) -> str:
    """Короткая подпись для кнопки в списке чеков: имя (или ID), магазин,
    username. Каждая часть обрезается ОТДЕЛЬНО (а не всё целиком), иначе
    при длинном ФИО магазин/username могли полностью "вытесняться" за
    пределы общего лимита длины текста кнопки."""
    reg = google_sheets.get_registration_by_telegram_id(r["telegram_id"])
    name = _truncate(reg["full_name"] if reg else f"ID {r['telegram_id']}", 12)
    shop = _truncate(reg.get("shop") if reg else "", 10)
    username = _truncate(f"@{r['username']}" if r.get("username") else "", 10)

    parts = [p for p in [name, shop, username] if p]
    return " · ".join(parts)


# ---------- История: календарь ----------

MONTHS_RU = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]


def _build_calendar(year: int, month: int, prefix: str = "cal"):
    """Строит инлайн-календарь на месяц. prefix различает, для какой
    именно цели используется календарь (история модератора, дата чека
    пользователя, начало/конец периода для отчёта) — у каждого свои
    callback_data и свои обработчики, чтобы они не пересекались."""
    builder = InlineKeyboardBuilder()

    builder.row(InlineKeyboardButton(text=f"{MONTHS_RU[month]} {year}", callback_data=f"{prefix}_ignore"))
    builder.row(*[
        InlineKeyboardButton(text=d, callback_data=f"{prefix}_ignore")
        for d in ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    ])

    for week in calendar.monthcalendar(year, month):
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(text=" ", callback_data=f"{prefix}_ignore"))
            else:
                row.append(InlineKeyboardButton(
                    text=str(day),
                    callback_data=f"{prefix}_day:{year:04d}-{month:02d}-{day:02d}",
                ))
        builder.row(*row)

    prev_year, prev_month = (year - 1, 12) if month == 1 else (year, month - 1)
    next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)
    builder.row(
        InlineKeyboardButton(text="◀️", callback_data=f"{prefix}_nav:{prev_year:04d}-{prev_month:02d}"),
        InlineKeyboardButton(text="▶️", callback_data=f"{prefix}_nav:{next_year:04d}-{next_month:02d}"),
    )
    return builder.as_markup()


@dp.message(F.text == ADMIN_BTN_HISTORY, IsModerator())
async def admin_history_start(message: Message):
    today = google_sheets.moscow_today()
    await message.answer("Выберите дату:", reply_markup=_build_calendar(today.year, today.month, prefix="cal"))


@dp.callback_query(F.data == "cal_ignore", IsModerator())
async def admin_calendar_ignore(call: CallbackQuery):
    await call.answer()


@dp.callback_query(F.data.startswith("cal_nav:"), IsModerator())
async def admin_calendar_nav(call: CallbackQuery):
    _, ym = call.data.split(":", 1)
    year, month = map(int, ym.split("-"))
    await call.message.edit_reply_markup(reply_markup=_build_calendar(year, month, prefix="cal"))
    await call.answer()


async def _send_day_receipt_list(message: Message, date_str: str):
    """Показывает список чеков за дату — используется и при выборе даты в
    календаре, и как "Назад" из карточки чека, и после удаления чека."""
    try:
        receipts = google_sheets.get_receipts_by_date(date_str)
    except Exception:
        logger.exception("Не удалось получить чеки за дату")
        await message.answer("Не получилось прочитать данные из таблицы, попробуйте ещё раз.")
        return

    if not receipts:
        await message.answer(f"На {date_str} чеков нет.")
        return

    builder = InlineKeyboardBuilder()
    for r in receipts:
        label = _format_receipt_label(r)
        time_part = r["date"][11:16] if len(r["date"]) >= 16 else r["date"]
        builder.row(InlineKeyboardButton(
            text=f"{time_part} — {label} ({r['status']})",
            callback_data=f"hist_view:{r['row']}",
        ))
    await message.answer(f"Чеки за {date_str}:", reply_markup=builder.as_markup())


@dp.callback_query(F.data.startswith("cal_day:"), IsModerator())
async def admin_calendar_pick_day(call: CallbackQuery):
    _, date_str = call.data.split(":", 1)
    await _send_day_receipt_list(call.message, date_str)
    await call.answer()


@dp.callback_query(F.data.startswith("hist_view:"), IsModerator())
async def admin_history_view(call: CallbackQuery):
    row = int(call.data.split(":", 1)[1])
    receipt = google_sheets.get_receipt_by_row(row)
    if not receipt:
        await call.answer("Не найдено", show_alert=True)
        return
    await _send_receipt_card(call.message, receipt, with_history_buttons=True, viewer_id=call.from_user.id)
    await call.answer()


@dp.callback_query(F.data.startswith("hist_delete:"), IsModerator())
async def admin_history_delete_confirm(call: CallbackQuery):
    row = int(call.data.split(":", 1)[1])
    kb = InlineKeyboardBuilder()
    kb.row(
        InlineKeyboardButton(text="✅ Да, удалить", callback_data=f"hist_delete_yes:{row}"),
        InlineKeyboardButton(text="❌ Отмена", callback_data="hist_delete_no"),
    )
    await call.message.answer(
        "Точно удалить этот чек из таблицы? Действие необратимо.",
        reply_markup=kb.as_markup(),
    )
    await call.answer()


@dp.callback_query(F.data == "hist_delete_no", IsModerator())
async def admin_history_delete_cancel(call: CallbackQuery):
    await call.message.answer("Отменено, чек не удалён.")
    await call.answer()


@dp.callback_query(F.data.startswith("hist_delete_yes:"), IsModerator())
async def admin_history_delete_do(call: CallbackQuery):
    row = int(call.data.split(":", 1)[1])
    receipt = google_sheets.get_receipt_by_row(row)
    if not receipt:
        await call.answer("Уже удалено или не найдено.", show_alert=True)
        return

    date_str = receipt["date"][:10] if receipt.get("date") else ""

    try:
        google_sheets.mark_receipt_deleted(row)
    except Exception:
        logger.exception("Не удалось пометить чек как удалённый")
        await call.message.answer("Не получилось обновить таблицу, попробуйте ещё раз.")
        await call.answer()
        return

    await call.message.answer(
        "Чек помечен как удалённый и больше не будет виден в списках. "
        "Сама запись и фото остаются в таблице/на диске."
    )
    if date_str:
        await _send_day_receipt_list(call.message, date_str)
    await call.answer()


# ---------- Модерация чеков ----------

@dp.message(F.text == ADMIN_BTN_MODERATION, IsModerator())
async def admin_moderation_menu(message: Message):
    kb = InlineKeyboardBuilder()
    kb.row(InlineKeyboardButton(text="📋 Список на модерации", callback_data="moderation_list"))
    kb.row(InlineKeyboardButton(text="🤖 Автоматическая модерация", callback_data="moderation_auto"))
    await message.answer("Что открыть?", reply_markup=kb.as_markup())


async def _show_moderation_list(message: Message):
    try:
        receipts = google_sheets.get_pending_receipts()
    except Exception:
        logger.exception("Не удалось получить чеки на модерации")
        await message.answer("Не получилось прочитать данные из таблицы, попробуйте ещё раз.")
        return

    if not receipts:
        await message.answer("Чеков на модерации нет.")
        return

    builder = InlineKeyboardBuilder()
    for r in receipts:
        label = _format_receipt_label(r)
        builder.row(InlineKeyboardButton(text=f"{r['date']} — {label}", callback_data=f"mod_view:{r['row']}"))
    await message.answer("Чеки на модерации:", reply_markup=builder.as_markup())


@dp.callback_query(F.data == "moderation_list", IsModerator())
async def admin_moderation_list_cb(call: CallbackQuery):
    await call.answer()
    await _show_moderation_list(call.message)


@dp.callback_query(F.data.startswith("mod_view:"), IsModerator())
async def admin_moderation_view(call: CallbackQuery):
    row = int(call.data.split(":", 1)[1])
    receipt = google_sheets.get_receipt_by_row(row)
    if not receipt:
        await call.answer("Не найдено", show_alert=True)
        return
    await _send_receipt_card(call.message, receipt, with_moderation_buttons=True, viewer_id=call.from_user.id)
    await call.answer()


async def _reject_receipt(row: int, reason_text: str):
    """Ставит чеку статус 'отклонён' и уведомляет пользователя причиной.
    Возвращает (True, notified, None) при успешном обновлении статуса
    (notified — удалось ли отправить уведомление пользователю), или
    (False, False, текст_ошибки), если не получилось даже обновить
    статус."""
    receipt = google_sheets.get_receipt_by_row(row)
    if not receipt:
        return False, False, "Не найдено."

    try:
        google_sheets.update_receipt_status(row, google_sheets.STATUS_REJECTED, comment=reason_text)
    except Exception:
        logger.exception("Не удалось обновить статус чека")
        return False, False, "Не получилось обновить таблицу, попробуйте ещё раз."

    notified = True
    try:
        await bot.send_message(int(receipt["telegram_id"]), reason_text)
    except Exception:
        logger.exception(
            "Не удалось уведомить пользователя %s об отклонении чека (строка %s)",
            receipt["telegram_id"], row,
        )
        notified = False

    return True, notified, None


@dp.callback_query(F.data.startswith("mod_reject_photo:"), IsModerator())
async def admin_moderation_reject_photo(call: CallbackQuery):
    row = int(call.data.split(":", 1)[1])
    ok, notified, error = await _reject_receipt(row, google_sheets.get_text("reject_message"))
    if not ok:
        await call.answer(error, show_alert=True)
        return
    if notified:
        await call.message.answer("Чек отклонён (быстрое отклонение), пользователь уведомлён.")
    else:
        await call.message.answer(
            "Чек отклонён (быстрое отклонение), но отправить уведомление "
            "пользователю не удалось (возможно, он заблокировал бота или "
            "ни разу не запускал его)."
        )
    await call.answer()


@dp.callback_query(F.data.startswith("mod_reject_custom:"), IsModerator())
async def admin_moderation_reject_custom_start(call: CallbackQuery, state: FSMContext):
    row = int(call.data.split(":", 1)[1])
    await state.set_state(AdminForm.waiting_reject_reason)
    await state.update_data(moderation_row=row)
    await call.message.answer(
        "Введите причину отклонения одним сообщением — она будет отправлена пользователю."
    )
    await call.answer()


@dp.message(StateFilter(AdminForm.waiting_reject_reason), IsModerator())
async def admin_moderation_reject_custom_finish(message: Message, state: FSMContext):
    data = await state.get_data()
    row = data.get("moderation_row")

    ok, notified, error = await _reject_receipt(row, message.text)
    await state.clear()

    menu = _menu_for(message.from_user.id)
    if not ok:
        await message.answer(error, reply_markup=menu)
        return
    if notified:
        await message.answer("Чек отклонён, пользователь уведомлён.", reply_markup=menu)
    else:
        await message.answer(
            "Чек отклонён, но отправить уведомление пользователю не "
            "удалось (возможно, он заблокировал бота или ни разу не "
            "запускал его).",
            reply_markup=menu,
        )


@dp.callback_query(F.data.startswith("mod_accept:"), IsModerator())
async def admin_moderation_accept_start(call: CallbackQuery, state: FSMContext):
    row = int(call.data.split(":", 1)[1])
    await state.set_state(AdminForm.accept_coupon_choice)
    await state.update_data(moderation_row=row)
    kb = InlineKeyboardBuilder()
    kb.row(
        InlineKeyboardButton(text="✅ Да", callback_data="accept_coupon_yes"),
        InlineKeyboardButton(text="❌ Нет", callback_data="accept_coupon_no"),
    )
    await call.message.answer("Чек принимается. Отправить купон?", reply_markup=kb.as_markup())
    await call.answer()


@dp.callback_query(F.data == "accept_coupon_yes", IsModerator(), StateFilter(AdminForm.accept_coupon_choice))
async def admin_accept_coupon_yes(call: CallbackQuery, state: FSMContext):
    await state.set_state(AdminForm.waiting_coupon)
    await call.message.answer("Введите номер купона одним сообщением — я отправлю его пользователю.")
    await call.answer()


@dp.callback_query(F.data == "accept_coupon_no", IsModerator(), StateFilter(AdminForm.accept_coupon_choice))
async def admin_accept_coupon_no(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    row = data.get("moderation_row")
    await state.clear()
    await call.answer()

    receipt = google_sheets.get_receipt_by_row(row)
    if not receipt:
        await call.message.answer("Не удалось найти этот чек (возможно, таблица изменилась).")
        return

    try:
        google_sheets.update_receipt_status(row, google_sheets.STATUS_ACCEPTED)
    except Exception:
        logger.exception("Не удалось обновить статус чека")
        await call.message.answer("Не получилось обновить таблицу, попробуйте ещё раз.")
        return

    menu = _menu_for(call.from_user.id)
    try:
        text = google_sheets.get_text("accept_no_coupon_message")
        await bot.send_message(int(receipt["telegram_id"]), text)
        await call.message.answer("Чек принят без купона, пользователь уведомлён.", reply_markup=menu)
    except Exception:
        logger.exception("Не удалось уведомить пользователя %s о принятии чека без купона (строка %s)", receipt["telegram_id"], row)
        await call.message.answer(
            "Чек принят без купона, но уведомить пользователя не удалось "
            "(возможно, он заблокировал бота).",
            reply_markup=menu,
        )


@dp.callback_query(F.data.in_({"accept_coupon_yes", "accept_coupon_no"}))
async def admin_accept_coupon_stale(call: CallbackQuery):
    await call.answer("Эта кнопка уже не активна.", show_alert=True)


@dp.message(StateFilter(AdminForm.waiting_coupon), IsModerator())
async def admin_moderation_accept_finish(message: Message, state: FSMContext):
    data = await state.get_data()
    row = data.get("moderation_row")
    coupon = message.text

    receipt = google_sheets.get_receipt_by_row(row)
    if not receipt:
        await message.answer("Не удалось найти этот чек (возможно, таблица изменилась).")
        await state.clear()
        return

    try:
        google_sheets.update_receipt_status(row, google_sheets.STATUS_ACCEPTED, coupon=coupon)
    except Exception:
        logger.exception("Не удалось обновить статус чека")
        await message.answer("Не получилось обновить таблицу, попробуйте ещё раз.")
        await state.clear()
        return

    try:
        accept_text_template = google_sheets.get_text("accept_message")
        try:
            accept_text = accept_text_template.format(
                coupon=coupon,
                upd_number=receipt.get("upd_number", ""),
                receipt_date=receipt.get("receipt_date", ""),
            )
        except Exception:
            logger.exception("Не удалось подставить данные в шаблон сообщения, использую текст по умолчанию")
            accept_text = (
                f"Ваш купон № {coupon} по УПД № {receipt.get('upd_number', '')} "
                f"от {receipt.get('receipt_date', '')}"
            )

        await bot.send_message(int(receipt["telegram_id"]), accept_text)
        await message.answer("Купон отправлен пользователю, статус обновлён.", reply_markup=_menu_for(message.from_user.id))
    except Exception:
        logger.exception("Не удалось отправить купон пользователю %s (строка %s)", receipt["telegram_id"], row)
        await message.answer(
            "Статус обновлён, но отправить сообщение пользователю не удалось "
            "(возможно, он заблокировал бота).",
            reply_markup=_menu_for(message.from_user.id),
        )

    await state.clear()


# ---------- Управление модераторами (только владельцы из ADMIN_IDS) ----------

@dp.message(F.text == ADMIN_BTN_MODERATORS, IsOwner())
async def admin_moderators_menu(message: Message):
    try:
        moderators = google_sheets.get_moderators()
    except Exception:
        logger.exception("Не удалось получить список модераторов")
        await message.answer("Не получилось прочитать данные из таблицы, попробуйте ещё раз.")
        return

    builder = InlineKeyboardBuilder()
    for m in moderators:
        label = m["username"] or m["telegram_id"]
        builder.row(InlineKeyboardButton(text=f"❌ Удалить: {label}", callback_data=f"modrole_del:{m['row']}"))
    builder.row(InlineKeyboardButton(text="➕ Добавить модератора", callback_data="modrole_add"))

    if moderators:
        listing = "\n".join(
            f"- {m['telegram_id']}" + (f" (@{m['username']})" if m["username"] else "")
            for m in moderators
        )
    else:
        listing = "пока никого нет."

    await message.answer(f"Текущие модераторы:\n{listing}", reply_markup=builder.as_markup())


@dp.callback_query(F.data == "modrole_add", IsOwner())
async def admin_moderator_add_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(AdminForm.waiting_moderator_id)
    await call.message.answer(
        "Пришлите Telegram ID нового модератора одним сообщением (число). "
        "Попросите его узнать свой ID через @userinfobot и прислать вам."
    )
    await call.answer()


@dp.message(StateFilter(AdminForm.waiting_moderator_id), IsOwner())
async def admin_moderator_add_finish(message: Message, state: FSMContext):
    text = (message.text or "").strip()
    if not text.isdigit():
        await message.answer("Это не похоже на Telegram ID — пришлите число, например 123456789.")
        return

    new_id = int(text)

    # Username нельзя узнать из одного только введённого числа — пробуем
    # спросить у Telegram напрямую. Это сработает, только если новый
    # модератор уже хотя бы раз писал этому боту раньше (иначе Telegram
    # не отдаёт информацию о чате) — если нет, username останется пустым,
    # это не критично: права проверяются по Telegram ID, а не по username.
    username = ""
    try:
        chat = await bot.get_chat(new_id)
        username = chat.username or ""
    except Exception:
        logger.info(
            "Не удалось получить username для %s (скорее всего, он ещё не писал боту) — "
            "сохраняю без username",
            new_id,
        )

    try:
        google_sheets.add_moderator(new_id, username=username, added_by=message.from_user.id)
    except Exception:
        logger.exception("Не удалось добавить модератора")
        await message.answer("Не получилось сохранить в таблицу, попробуйте ещё раз.")
        await state.clear()
        return

    await state.clear()
    await message.answer(f"Готово, {new_id} теперь модератор.", reply_markup=owner_menu)

    try:
        await bot.send_message(new_id, "Вас назначили модератором бота. Отправьте /admin, чтобы открыть панель.")
    except Exception:
        logger.exception("Не удалось уведомить нового модератора (возможно, он ещё не запускал бота)")


@dp.callback_query(F.data.startswith("modrole_del:"), IsOwner())
async def admin_moderator_delete(call: CallbackQuery):
    row = int(call.data.split(":", 1)[1])
    try:
        google_sheets.remove_moderator(row)
        await call.message.answer("Модератор удалён.")
    except Exception:
        logger.exception("Не удалось удалить модератора")
        await call.message.answer("Не получилось удалить, попробуйте ещё раз.")
    await call.answer()


# ---------- Редактирование текстов-автоответов (только владельцы) ----------

def _texts_menu_markup():
    builder = InlineKeyboardBuilder()
    for key, label in google_sheets.TEXT_LABELS.items():
        builder.row(InlineKeyboardButton(text=label, callback_data=f"text_edit:{key}"))
    return builder.as_markup()


@dp.message(F.text == ADMIN_BTN_TEXTS, IsOwner())
async def admin_texts_menu(message: Message):
    await message.answer("Какой текст изменить?", reply_markup=_texts_menu_markup())


@dp.callback_query(F.data.startswith("text_edit:"), IsOwner())
async def admin_text_edit_start(call: CallbackQuery, state: FSMContext):
    key = call.data.split(":", 1)[1]
    label = google_sheets.TEXT_LABELS.get(key, key)

    try:
        current = google_sheets.get_text(key)
    except Exception:
        logger.exception("Не удалось прочитать текущий текст")
        await call.message.answer("Не получилось прочитать данные из таблицы, попробуйте ещё раз.")
        await call.answer()
        return

    await state.set_state(AdminForm.waiting_text_value)
    await state.update_data(text_key=key)

    kb = InlineKeyboardBuilder()
    kb.row(InlineKeyboardButton(text="❌ Отмена", callback_data="text_edit_cancel"))
    await call.message.answer(
        f"«{label}»\n\nТекущий текст:\n{current}\n\nПришлите новый текст одним сообщением.",
        reply_markup=kb.as_markup(),
    )
    await call.answer()


@dp.callback_query(F.data == "text_edit_cancel", IsOwner(), StateFilter(AdminForm.waiting_text_value))
async def admin_text_edit_cancel(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.answer()
    await call.message.answer("Изменение текста отменено.", reply_markup=_texts_menu_markup())


@dp.callback_query(F.data == "text_edit_cancel")
async def admin_text_edit_cancel_stale(call: CallbackQuery):
    await call.answer("Эта кнопка уже не активна.", show_alert=True)


@dp.message(StateFilter(AdminForm.waiting_text_value), IsOwner())
async def admin_text_edit_finish(message: Message, state: FSMContext):
    data = await state.get_data()
    key = data.get("text_key")
    new_value = message.text

    try:
        google_sheets.set_text(key, new_value)
    except Exception:
        logger.exception("Не удалось сохранить текст")
        await message.answer("Не получилось сохранить в таблицу, попробуйте ещё раз.")
        await state.clear()
        return

    await state.clear()
    await message.answer("Готово, текст обновлён.", reply_markup=owner_menu)


# ---------- Импорт (акционные позиции / отчёт по продажам) ----------

@dp.message(F.text == ADMIN_BTN_IMPORT, IsOwner())
async def admin_import_menu(message: Message):
    kb = InlineKeyboardBuilder()
    kb.row(InlineKeyboardButton(text="📦 Акционные позиции", callback_data="import_promo"))
    kb.row(InlineKeyboardButton(text="📈 Отчёт по продажам", callback_data="import_sales"))
    kb.row(InlineKeyboardButton(text="🎟 Купоны", callback_data="import_coupons"))
    await message.answer(
        "Что загружаем? Все варианты — файл .xlsx, первая строка листа "
        "должна быть заголовком (её содержимое не важно, бот её "
        "пропускает и читает данные со 2-й строки).\n\n"
        "📦 Акционные позиции — колонки: Артикул, Наименование. "
        "Новая загрузка полностью заменяет предыдущий список.\n"
        "📈 Отчёт по продажам — колонки: ИНН, Наименование, Артикул, "
        "Количество, Цена, № УПД, Дата продажи. Новая загрузка "
        "ДОБАВЛЯЕТСЯ к уже имеющимся данным (ничего не заменяется), и "
        "каждая строка дополнительно помечается датой и временем именно "
        "этой загрузки.\n\n"
        "🎟 Купоны — колонки: №, Username, купон, номинал. После загрузки "
        "файла бот спросит подтверждение и разошлёт купоны пользователям "
        "личным сообщением (по Username) — ничего не сохраняется в "
        "Google Таблицу, это разовая рассылка.\n\n"
        "Если нужен пустой шаблон файла для заполнения — загляните в "
        f"«{ADMIN_BTN_TEMPLATES}».",
        reply_markup=kb.as_markup(),
    )


@dp.callback_query(F.data == "import_promo", IsOwner())
async def admin_import_promo_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(AdminForm.waiting_promo_file)
    await call.message.answer(
        "Пришлите файл .xlsx со списком акционных позиций "
        "(колонки: Артикул, Наименование)."
    )
    await call.answer()


@dp.callback_query(F.data == "import_sales", IsOwner())
async def admin_import_sales_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(AdminForm.waiting_sales_file)
    await call.message.answer(
        "Пришлите файл .xlsx с отчётом по продажам (колонки: ИНН, "
        "Наименование, Артикул, Количество, Цена, № УПД, Дата продажи)."
    )
    await call.answer()


@dp.callback_query(F.data == "import_coupons", IsOwner())
async def admin_import_coupons_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(AdminForm.waiting_coupons_file)
    await call.message.answer(
        "Пришлите файл .xlsx со списком купонов (колонки: №, Username, "
        "купон, номинал)."
    )
    await call.answer()


def _xlsx_cell_to_str(cell) -> str:
    """Аккуратно превращает значение ячейки Excel в текст:
    - даты (date/datetime) -> ДД.ММ.ГГГГ, а не "2026-07-15 00:00:00";
    - "круглые" числа (10.0) -> "10", а не "10.0" (важно для ИНН/УПД/артикулов);
    - остальное -> обычный str().strip()."""
    if cell is None:
        return ""
    if isinstance(cell, datetime.datetime):
        return cell.strftime("%d.%m.%Y")
    if isinstance(cell, date):
        return cell.strftime("%d.%m.%Y")
    if isinstance(cell, float):
        if cell.is_integer():
            return str(int(cell))
        return str(cell).strip()
    return str(cell).strip()


def _read_xlsx_rows(file_bytes: io.BytesIO, min_columns: int):
    """Читает .xlsx: первая строка — заголовок (пропускается), из
    остальных строк берёт первые min_columns колонок текстом. Полностью
    пустые строки пропускаются."""
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(cell in (None, "") for cell in row):
            continue
        values = [_xlsx_cell_to_str(cell) for cell in row[:min_columns]]
        if len(values) < min_columns:
            values += [""] * (min_columns - len(values))
        rows.append(values)
    return rows


def _looks_like_product_name(value: str) -> bool:
    """Грубая эвристика: похоже ли значение скорее на наименование
    товара, чем на короткий код артикула (длинная строка и/или
    несколько слов с кириллицей). Используется только для
    предупреждения администратора, если он, похоже, перепутал местами
    колонки «Артикул» и «Наименование» при подготовке файла — сам
    импорт при этом всё равно проходит как есть."""
    v = value.strip()
    if not v:
        return False
    if len(v) > 35:
        return True
    return bool(re.search(r"[а-яА-ЯёЁ]", v)) and len(v.split()) >= 3


def _column_swap_warning(rows, article_index: int) -> str:
    """Если у большинства строк значение в колонке артикула похоже на
    наименование товара, возвращает предупреждающий текст (иначе — "")."""
    if not rows:
        return ""
    suspicious = sum(1 for row in rows if _looks_like_product_name(row[article_index]))
    if suspicious / len(rows) <= 0.3:
        return ""
    return (
        "\n\n⚠️ Внимание: у большинства строк в колонке «Артикул» — "
        "длинные текстовые описания, а не короткие коды. Похоже, в файле "
        "перепутаны местами колонки «Артикул» и «Наименование» "
        "(наименование должно идти ПОСЛЕ артикула). Проверьте файл и "
        "загрузите заново, если это так — импорт полностью заменяет "
        "предыдущий список, так что это безопасно повторить."
    )


async def _download_document(message: Message) -> io.BytesIO:
    buf = io.BytesIO()
    await bot.download(message.document.file_id, destination=buf)
    buf.seek(0)
    return buf


@dp.message(StateFilter(AdminForm.waiting_promo_file), F.document)
async def admin_import_promo_finish(message: Message, state: FSMContext):
    if not message.document.file_name.lower().endswith((".xlsx", ".xlsm")):
        await message.answer("Нужен файл в формате .xlsx. Пришлите файл ещё раз.")
        return

    try:
        buf = await _download_document(message)
        rows = _read_xlsx_rows(buf, min_columns=2)
        google_sheets.import_promo_items(rows)
    except Exception:
        logger.exception("Не удалось импортировать акционные позиции")
        await message.answer(
            "Не получилось прочитать файл. Проверьте формат и попробуйте ещё раз.",
        )
        return

    await state.clear()
    warning = _column_swap_warning(rows, article_index=0)
    await message.answer(f"Готово, загружено позиций: {len(rows)}.{warning}", reply_markup=owner_menu)


@dp.message(StateFilter(AdminForm.waiting_promo_file))
async def wrong_content_for_promo_file(message: Message):
    await message.answer("Пожалуйста, пришлите файл .xlsx (документом).")


@dp.message(StateFilter(AdminForm.waiting_sales_file), F.document)
async def admin_import_sales_finish(message: Message, state: FSMContext):
    if not message.document.file_name.lower().endswith((".xlsx", ".xlsm")):
        await message.answer("Нужен файл в формате .xlsx. Пришлите файл ещё раз.")
        return

    try:
        buf = await _download_document(message)
        rows = _read_xlsx_rows(buf, min_columns=7)
        google_sheets.import_sales_report(rows)
    except Exception:
        logger.exception("Не удалось импортировать отчёт по продажам")
        await message.answer(
            "Не получилось прочитать файл. Проверьте формат и попробуйте ещё раз.",
        )
        return

    await state.clear()
    warning = _column_swap_warning(rows, article_index=2)
    await message.answer(
        f"Готово, добавлено строк: {len(rows)} (дата загрузки проставлена автоматически).{warning}",
        reply_markup=owner_menu,
    )


@dp.message(StateFilter(AdminForm.waiting_sales_file))
async def wrong_content_for_sales_file(message: Message):
    await message.answer("Пожалуйста, пришлите файл .xlsx (документом).")


@dp.message(StateFilter(AdminForm.waiting_coupons_file), F.document)
async def admin_import_coupons_finish(message: Message, state: FSMContext):
    if not message.document.file_name.lower().endswith((".xlsx", ".xlsm")):
        await message.answer("Нужен файл в формате .xlsx. Пришлите файл ещё раз.")
        return

    try:
        buf = await _download_document(message)
        raw_rows = _read_xlsx_rows(buf, min_columns=4)
    except Exception:
        logger.exception("Не удалось прочитать файл с купонами")
        await message.answer("Не получилось прочитать файл. Проверьте формат и попробуйте ещё раз.")
        return

    # row[0] — "№" (просто для удобства заполнения, боту не нужен).
    coupon_rows = [
        {"username": row[1].strip(), "coupon": row[2].strip(), "nominal": row[3].strip()}
        for row in raw_rows
        if row[1].strip() and row[2].strip()
    ]

    if not coupon_rows:
        await state.clear()
        await message.answer(
            "В файле не нашлось ни одной строки с заполненными Username и "
            "купоном. Проверьте файл и загрузите заново.",
            reply_markup=owner_menu,
        )
        return

    await state.set_state(AdminForm.coupons_confirm)
    await state.update_data(coupon_rows=coupon_rows)

    skipped = len(raw_rows) - len(coupon_rows)
    skipped_note = f"\n(строк пропущено из-за пустых Username/купона: {skipped})" if skipped else ""

    kb = InlineKeyboardBuilder()
    kb.row(
        InlineKeyboardButton(text="✅ Да", callback_data="coupons_send_yes"),
        InlineKeyboardButton(text="❌ Нет", callback_data="coupons_send_no"),
    )
    await message.answer(
        f"В файле {len(coupon_rows)} купон(ов) для рассылки.{skipped_note}\n\n"
        "Отправить купоны пользователям?",
        reply_markup=kb.as_markup(),
    )


@dp.message(StateFilter(AdminForm.waiting_coupons_file))
async def wrong_content_for_coupons_file(message: Message):
    await message.answer("Пожалуйста, пришлите файл .xlsx (документом).")


@dp.callback_query(F.data == "coupons_send_yes", IsOwner(), StateFilter(AdminForm.coupons_confirm))
async def admin_coupons_send_yes(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    coupon_rows = data.get("coupon_rows", [])
    await state.clear()
    await call.answer()
    await call.message.answer("Рассылаю купоны, секунду...")

    message_template = google_sheets.get_text("coupon_import_message")

    sent = 0
    not_found = []
    failed = []
    log_rows = []

    for row in coupon_rows:
        username = row["username"]
        telegram_id = google_sheets.get_telegram_id_by_username(username)
        if not telegram_id:
            not_found.append(username)
            log_rows.append([username, "", row["coupon"], row["nominal"], "username не найден"])
            continue

        try:
            text = message_template.format(coupon=row["coupon"], nominal=row["nominal"])
        except Exception:
            text = f"Ваш купон для OZON: {row['coupon']}, с номиналом {row['nominal']}"

        try:
            await bot.send_message(int(telegram_id), text)
            sent += 1
            log_rows.append([username, telegram_id, row["coupon"], row["nominal"], "отправлено"])
        except Exception:
            logger.exception("Не удалось отправить купон пользователю @%s", username)
            failed.append(username)
            log_rows.append([username, telegram_id, row["coupon"], row["nominal"], "ошибка отправки"])

    try:
        google_sheets.append_coupons_log(log_rows)
    except Exception:
        logger.exception("Не удалось записать лог рассылки купонов в Google Таблицу")

    summary = [f"Готово. Отправлено: {sent} из {len(coupon_rows)}."]
    if not_found:
        summary.append(
            "Не найден Username (нет такой регистрации): " + ", ".join(f"@{u}" for u in not_found[:20])
            + ("…" if len(not_found) > 20 else "")
        )
    if failed:
        summary.append(
            "Не удалось отправить сообщение (например, пользователь не "
            "запускал бота): " + ", ".join(f"@{u}" for u in failed[:20])
            + ("…" if len(failed) > 20 else "")
        )

    await call.message.answer("\n\n".join(summary), reply_markup=owner_menu)


@dp.callback_query(F.data == "coupons_send_no", IsOwner(), StateFilter(AdminForm.coupons_confirm))
async def admin_coupons_send_no(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.answer()
    await call.message.answer("Хорошо, рассылка отменена.", reply_markup=owner_menu)


@dp.callback_query(F.data.in_({"coupons_send_yes", "coupons_send_no"}))
async def admin_coupons_send_stale(call: CallbackQuery):
    await call.answer("Эта кнопка уже не активна.", show_alert=True)


# ---------- Шаблоны для импорта ----------

TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")

# callback_data -> (имя файла в папке templates/, подпись к документу)
TEMPLATE_FILES = {
    "template_promo": (
        "promo_items_template.xlsx",
        "Шаблон для импорта акционных позиций. Колонки: Артикул, "
        "Наименование (именно в этом порядке — первая колонка короткий "
        "код артикула, вторая текстовое наименование товара).",
    ),
    "template_sales": (
        "sales_report_template.xlsx",
        "Шаблон для импорта отчёта по продажам. Колонки: ИНН, "
        "Наименование, Артикул, Количество, Цена, № УПД, Дата продажи.",
    ),
    "template_coupons": (
        "coupons_template.xlsx",
        "Шаблон для рассылки купонов. Колонки: №, Username, купон, "
        "номинал.",
    ),
}


@dp.message(F.text == ADMIN_BTN_TEMPLATES, IsOwner())
async def admin_templates_menu(message: Message):
    kb = InlineKeyboardBuilder()
    kb.row(InlineKeyboardButton(text="📦 Акционные позиции", callback_data="template_promo"))
    kb.row(InlineKeyboardButton(text="📈 Отчёт по продажам", callback_data="template_sales"))
    kb.row(InlineKeyboardButton(text="🎟 Купоны", callback_data="template_coupons"))
    await message.answer("Какой шаблон для импорта прислать?", reply_markup=kb.as_markup())


@dp.callback_query(F.data.in_(TEMPLATE_FILES.keys()), IsOwner())
async def admin_templates_send(call: CallbackQuery):
    filename, caption = TEMPLATE_FILES[call.data]
    path = os.path.join(TEMPLATES_DIR, filename)
    await call.answer()

    try:
        with open(path, "rb") as f:
            data = f.read()
    except OSError:
        logger.exception("Не найден файл шаблона: %s", path)
        await call.message.answer(
            "Не получилось найти файл шаблона на сервере "
            f"(ожидался файл templates/{filename} рядом с bot.py). "
            "Убедитесь, что папка «templates» загружена вместе с "
            "остальным кодом бота при деплое."
        )
        return

    await call.message.answer_document(BufferedInputFile(data, filename=filename), caption=caption)


# ---------- Отчёт в Excel ----------

def _autosize(ws, headers):
    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(header) + 4)


def _build_report_workbook() -> io.BytesIO:
    registrations = google_sheets.get_all_registrations()
    reg_by_id = {reg["telegram_id"]: reg for reg in registrations}

    wb = openpyxl.Workbook()

    ws_receipts = wb.active
    ws_receipts.title = "Чеки"
    receipt_headers = [
        "Дата загрузки", "№ УПД", "Дата чека", "Дата регистрации", "Telegram ID", "Username",
        "ФИО", "Магазин", "ИНН магазина", "Телефон", "Статус", "Купон", "Комментарий", "Удалён",
    ]
    ws_receipts.append(receipt_headers)
    for r in google_sheets.get_receipts():
        reg = reg_by_id.get(r["telegram_id"])
        ws_receipts.append([
            r["date"],
            r["upd_number"],
            r["receipt_date"],
            reg["date"] if reg else "",
            r["telegram_id"],
            r["username"],
            reg["full_name"] if reg else "",
            reg["shop"] if reg else "",
            reg["inn"] if reg else "",
            reg["phone"] if reg else "",
            r["status"],
            r["coupon"],
            r["comment"],
            "да" if r["deleted"] else "",
        ])
    _autosize(ws_receipts, receipt_headers)

    ws_reg = wb.create_sheet("Регистрации")
    reg_headers = ["Дата регистрации", "ФИО", "Магазин", "ИНН магазина", "Телефон", "Telegram ID", "Username"]
    ws_reg.append(reg_headers)
    for reg in registrations:
        ws_reg.append([reg["date"], reg["full_name"], reg["shop"], reg["inn"], reg["phone"], reg["telegram_id"], reg["username"]])
    _autosize(ws_reg, reg_headers)

    ws_promo = wb.create_sheet("Акционные товары")
    promo_headers = ["Артикул", "Наименование"]
    ws_promo.append(promo_headers)
    for item in google_sheets.get_promo_items():
        ws_promo.append([item["article"], item["name"]])
    _autosize(ws_promo, promo_headers)

    ws_coupons = wb.create_sheet("Купоны")
    coupons_headers = ["Дата отправки", "Username", "Telegram ID", "Купон", "Номинал", "Статус"]
    ws_coupons.append(coupons_headers)
    for c in google_sheets.get_coupons_log():
        ws_coupons.append([
            c["sent_date"], c["username"], c["telegram_id"], c["coupon"], c["nominal"], c["status"],
        ])
    _autosize(ws_coupons, coupons_headers)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@dp.message(F.text == ADMIN_BTN_REPORT, IsModerator())
async def admin_reports_menu(message: Message):
    kb = InlineKeyboardBuilder()
    kb.row(InlineKeyboardButton(text="📄 Общий отчёт", callback_data="report_general"))
    kb.row(InlineKeyboardButton(text="🔎 Отчёт по продажам", callback_data="report_sales_match"))
    kb.row(InlineKeyboardButton(text="👤 Отчёт по пользователям", callback_data="report_by_users"))
    await message.answer("Какой отчёт сформировать?", reply_markup=kb.as_markup())


@dp.callback_query(F.data == "report_general", IsModerator())
async def admin_report_general(call: CallbackQuery):
    await call.answer()
    await call.message.answer("Формирую отчёт, секунду...")
    try:
        buf = _build_report_workbook()
    except Exception:
        logger.exception("Не удалось сформировать отчёт")
        await call.message.answer("Не получилось сформировать отчёт, попробуйте позже.")
        return

    filename = f"report_{google_sheets.moscow_today().strftime('%Y%m%d')}.xlsx"
    await call.message.answer_document(BufferedInputFile(buf.read(), filename=filename))


# ---------- Отчёт по продажам (сверка чеков с отчётом по продажам) ----------

_ISO_DATE_FORMATS = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d")


def _parse_date_loose(text: str):
    """Разбирает дату в разных написаниях в объект date. Возвращает None,
    если не получилось. Понимает:
    - ДД.ММ.ГГГГ (а также с "/" или "-" вместо точек, и 2-значный год);
    - "ГГГГ-ММ-ДД" и "ГГГГ-ММ-ДД ЧЧ:ММ:СС" — так выглядит дата, если её
      когда-то записали в Google Таблицу как str(datetime) (например, из
      старых импортов отчёта по продажам, сделанных до того, как чтение
      .xlsx-дат было исправлено) — чтобы уже загруженные ранее строки с
      таким "испорченным" форматом тоже подхватывались сверкой, а не
      требовали повторной загрузки файла."""
    if not text:
        return None
    text = text.strip()

    for fmt in _ISO_DATE_FORMATS:
        try:
            return datetime.datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    normalized = re.sub(r"[/\-]", ".", text)
    parts = normalized.split(".")
    if len(parts) != 3:
        return None
    day_s, month_s, year_s = parts
    if len(year_s) == 2:
        year_s = "20" + year_s
    try:
        return date(int(year_s), int(month_s), int(day_s))
    except ValueError:
        return None


@dp.callback_query(F.data == "report_sales_match", IsModerator())
async def admin_report_sales_match_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(AdminForm.picking_period)
    await state.update_data(period_purpose="sales_match")
    today = google_sheets.moscow_today()
    await call.message.answer(
        "Сверка чеков с отчётом по продажам.\n\n"
        "Выберите дату НАЧАЛА периода в календаре ниже (сверка идёт по "
        "дате чека). Бот сверит чеки с датой чека в этом периоде (у "
        "которых указаны ИНН, номер УПД и дата) с загруженным отчётом по "
        "продажам: совпадением считается одинаковые ИНН + № УПД + дата "
        "продажи, и хотя бы один товар по этому УПД — из списка акционных "
        "позиций.",
        reply_markup=_build_calendar(today.year, today.month, prefix="pstart"),
    )
    await call.answer()


@dp.callback_query(F.data == "report_by_users", IsModerator())
async def admin_report_by_users_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(AdminForm.picking_period)
    await state.update_data(period_purpose="user_report")
    today = google_sheets.moscow_today()
    await call.message.answer(
        "Отчёт по пользователям.\n\n"
        "Выберите дату НАЧАЛА периода в календаре ниже (по дате чека). "
        "Бот соберёт все чеки пользователей за этот период, сверит их с "
        "отчётом по продажам (по ИНН + № УПД + дате чека) и пришлёт "
        "сводку проданных позиций по каждому пользователю: ФИО, "
        "username, ИНН, магазин, артикул, наименование, количество и "
        "сумма (только по позициям из списка акционных товаров — как и "
        "в «🔎 Отчёт по продажам»).",
        reply_markup=_build_calendar(today.year, today.month, prefix="pstart"),
    )
    await call.answer()


def _build_sales_match_workbook(start_date: date, end_date: date):
    """Сверяет чеки (ИНН магазина + № УПД + дата чека) с отчётом по
    продажам (ИНН + № УПД + дата продажи), проверяя, что хотя бы один
    товар по этому УПД входит в список акционных позиций. Возвращает
    (буфер с Excel, число совпадений-строк, число чеков без совпадения,
    список номеров строк чеков, у которых нашлось хотя бы одно
    совпадение — используется для массовой смены статуса)."""
    promo_articles = {
        item["article"].strip() for item in google_sheets.get_promo_items() if item["article"].strip()
    }

    sales_index = {}
    # Индекс "только по № УПД" — нужен исключительно для диагностики:
    # если строка по ИНН+УПД+дате не нашлась, но строки с таким же № УПД
    # в отчёте есть, покажем админу, чем именно они отличаются (другой
    # ИНН, другая дата, дата не распозналась и т.п.), не заставляя лезть
    # в саму Google Таблицу и гадать.
    sales_by_upd = {}
    for s in google_sheets.get_sales_report():
        upd = google_sheets.normalize_upd(s["upd_number"])
        inn_s = s["inn"].strip()
        sale_date = _parse_date_loose(s["sale_date"])
        if upd:
            sales_by_upd.setdefault(upd, []).append((inn_s, sale_date, s["sale_date"]))
        if not sale_date:
            continue
        key = (inn_s, upd, sale_date)
        sales_index.setdefault(key, []).append(s)

    matched = []
    unmatched = []

    for r in google_sheets.get_receipts():
        # Сверяем ВСЕ чеки за период — независимо от статуса модерации
        # (на модерации / принят / отклонён): единственное исключение —
        # помеченные удалёнными.
        if r["deleted"]:
            continue

        receipt_date = _parse_date_loose(r["receipt_date"]) if r["receipt_date"] else None
        # Если дата чека не указана (например, старые чеки, загруженные
        # ещё до того, как это поле стало обязательным), для определения
        # попадания в период используем дату ЗАГРУЗКИ чека — чтобы такие
        # чеки не выпадали из выборки молча, а попали в "Без совпадений"
        # с понятной причиной.
        upload_date = _parse_date_loose(r["date"]) if r["date"] else None
        period_ref_date = receipt_date or upload_date

        if not period_ref_date or not (start_date <= period_ref_date <= end_date):
            continue

        reg = google_sheets.get_registration_by_telegram_id(r["telegram_id"])

        if not r["upd_number"] or not receipt_date:
            missing = []
            if not r["upd_number"]:
                missing.append("№ УПД")
            if not receipt_date:
                missing.append("дата чека")
            detail = ""
            if not receipt_date and upload_date:
                detail = "Попал в период по дате загрузки чека, т.к. дата чека не указана."
            unmatched.append((
                r, reg,
                f"у чека не заполнено: {', '.join(missing)} — сверить с отчётом по продажам нельзя",
                detail,
            ))
            continue

        inn = (reg.get("inn") or "").strip() if reg else ""
        upd_number = google_sheets.normalize_upd(r["upd_number"])

        if not inn:
            unmatched.append((r, reg, "у пользователя не указан ИНН магазина", ""))
            continue

        sale_lines = sales_index.get((inn, upd_number, receipt_date), [])
        promo_lines = [s for s in sale_lines if s["article"].strip() in promo_articles]

        if promo_lines:
            for s in promo_lines:
                matched.append((r, reg, s))
            continue

        if sale_lines:
            found_articles = sorted({s["article"].strip() for s in sale_lines if s["article"].strip()})
            detail = (
                f"В отчёте по этому № УПД указаны артикулы: {', '.join(found_articles) or '—'} "
                "— ни один из них не входит в список акционных позиций."
            )
            unmatched.append((r, reg, "товары по этому УПД не входят в акционные позиции", detail))
            continue

        candidates = sales_by_upd.get(upd_number, [])
        if candidates:
            details = []
            for cand_inn, cand_date, cand_raw in candidates[:3]:
                if cand_date:
                    cand_date_str = cand_date.strftime("%d.%m.%Y")
                else:
                    cand_date_str = f"не удалось разобрать дату ('{cand_raw}')"
                if cand_inn != inn:
                    details.append(f"в отчёте по этому УПД ИНН «{cand_inn or '—'}», у пользователя «{inn}»")
                elif cand_date != receipt_date:
                    details.append(f"в отчёте по этому УПД дата «{cand_date_str}», дата чека «{receipt_date.strftime('%d.%m.%Y')}»")
                else:
                    details.append("ИНН и дата совпадают, но строка не нашлась при поиске — сообщите разработчику")
            detail = "Похожие строки в отчёте по продажам есть, но не совпадают: " + "; ".join(details)
        else:
            detail = "В отчёте по продажам нет ни одной строки с таким № УПД вообще."

        unmatched.append((r, reg, "нет строки в отчёте по продажам с таким ИНН/№ УПД/датой", detail))

    wb = openpyxl.Workbook()

    ws_matched = wb.active
    ws_matched.title = "Совпадения"
    matched_headers = [
        "Дата чека", "№ УПД", "Telegram ID", "Username", "ФИО", "Магазин", "ИНН магазина", "Статус чека",
        "Артикул", "Наименование товара", "Количество", "Цена",
    ]
    ws_matched.append(matched_headers)
    for r, reg, s in matched:
        ws_matched.append([
            r["receipt_date"], r["upd_number"], r["telegram_id"], r["username"],
            reg["full_name"] if reg else "", reg["shop"] if reg else "", reg["inn"] if reg else "", r["status"],
            s["article"], s["name"], s["quantity"], s["price"],
        ])
    _autosize(ws_matched, matched_headers)

    ws_unmatched = wb.create_sheet("Без совпадений")
    unmatched_headers = [
        "Дата чека", "№ УПД", "Telegram ID", "Username", "ФИО", "Магазин", "ИНН магазина", "Статус чека",
        "Причина", "Подробности",
    ]
    ws_unmatched.append(unmatched_headers)
    for r, reg, reason, detail in unmatched:
        ws_unmatched.append([
            r["receipt_date"], r["upd_number"], r["telegram_id"], r["username"],
            reg["full_name"] if reg else "", reg["shop"] if reg else "", reg["inn"] if reg else "", r["status"],
            reason, detail,
        ])
    _autosize(ws_unmatched, unmatched_headers)
    ws_unmatched.column_dimensions[get_column_letter(len(unmatched_headers))].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    matched_rows = sorted({r["row"] for r, _reg, _s in matched})
    return buf, len(matched), len(unmatched), matched_rows


def _parse_number_loose(text):
    """Разбирает число (количество/цена) из текста — на случай, если в
    ячейке была запятая вместо точки как десятичный разделитель.
    Возвращает float или None, если не получилось."""
    if text is None:
        return None
    text = str(text).strip().replace(",", ".").replace(" ", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _build_user_report_workbook(start_date: date, end_date: date):
    """«Отчёт по пользователям»: за период собирает все чеки пользователей
    (ИНН + № УПД + дата чека), сверяет их с отчётом по продажам — как и
    "🔎 Отчёт по продажам", учитываются только позиции из списка
    акционных товаров ("📦 Акционные позиции" через "📥 Импорт") — и
    группирует результат по пользователю + артикулу + цене, суммируя
    количество (и, соответственно, сумму = кол-во × цена). Возвращает
    (буфер с Excel, число строк в сводке)."""
    promo_articles = {
        item["article"].strip() for item in google_sheets.get_promo_items() if item["article"].strip()
    }

    sales_index = {}
    for s in google_sheets.get_sales_report():
        sale_date = _parse_date_loose(s["sale_date"])
        if not sale_date:
            continue
        key = (s["inn"].strip(), google_sheets.normalize_upd(s["upd_number"]), sale_date)
        sales_index.setdefault(key, []).append(s)

    # (telegram_id, артикул, цена) -> накопленные данные по позиции
    groups = {}

    for r in google_sheets.get_receipts():
        if r["deleted"] or not r["upd_number"] or not r["receipt_date"]:
            continue

        receipt_date = _parse_date_loose(r["receipt_date"])
        if not receipt_date or not (start_date <= receipt_date <= end_date):
            continue

        reg = google_sheets.get_registration_by_telegram_id(r["telegram_id"])
        inn = (reg.get("inn") or "").strip() if reg else ""
        if not inn:
            continue

        sale_lines = sales_index.get((inn, google_sheets.normalize_upd(r["upd_number"]), receipt_date), [])
        promo_lines = [s for s in sale_lines if s["article"].strip() in promo_articles]

        for s in promo_lines:
            article = s["article"].strip()
            if not article:
                continue

            quantity = _parse_number_loose(s["quantity"])
            price = _parse_number_loose(s["price"])
            if quantity is None or price is None:
                continue

            group_key = (r["telegram_id"], article, price)
            group = groups.get(group_key)
            if group is None:
                group = {
                    "full_name": reg["full_name"] if reg else "",
                    "username": r["username"],
                    "inn": inn,
                    "shop": reg["shop"] if reg else "",
                    "article": article,
                    "name": s["name"],
                    "price": price,
                    "quantity": 0.0,
                }
                groups[group_key] = group
            group["quantity"] += quantity

    rows = sorted(groups.values(), key=lambda g: (g["full_name"], g["article"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт по пользователям"
    headers = [
        "№", "ФИО", "Username", "ИНН", "Магазин", "Артикул",
        "Наименование позиции", "Шт", "Цена", "Сумма",
    ]
    ws.append(headers)
    for i, g in enumerate(rows, start=1):
        quantity = g["quantity"]
        quantity_display = int(quantity) if quantity == int(quantity) else quantity
        amount = quantity * g["price"]
        ws.append([
            i, g["full_name"], g["username"], g["inn"], g["shop"],
            g["article"], g["name"], quantity_display, g["price"], amount,
        ])
    _autosize(ws, headers)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, len(rows)


# ---------- Автоматическая модерация (та же сверка + смена статуса) ----------

@dp.callback_query(F.data == "moderation_auto", IsModerator())
async def admin_moderation_auto_start(call: CallbackQuery, state: FSMContext):
    await state.set_state(AdminForm.picking_period)
    await state.update_data(period_purpose="auto_mod")
    today = google_sheets.moscow_today()
    await call.message.answer(
        "Автоматическая модерация.\n\n"
        "Выберите дату НАЧАЛА периода в календаре ниже (по дате чека). "
        "Бот сверит чеки за этот период с отчётом по продажам (так же, "
        "как «📊 Отчёты» -> «🔎 Отчёт по продажам»), пришлёт Excel и "
        "предложит сразу проставить статус «принят» найденным "
        "совпадениям.",
        reply_markup=_build_calendar(today.year, today.month, prefix="pstart"),
    )
    await call.answer()


# ---------- Общий календарь выбора периода (начало/конец) для обеих задач выше ----------

@dp.callback_query(F.data == "pstart_ignore")
async def period_start_calendar_ignore(call: CallbackQuery):
    await call.answer()


@dp.callback_query(F.data.startswith("pstart_nav:"), StateFilter(AdminForm.picking_period))
async def period_start_calendar_nav(call: CallbackQuery):
    _, ym = call.data.split(":", 1)
    year, month = map(int, ym.split("-"))
    await call.message.edit_reply_markup(reply_markup=_build_calendar(year, month, prefix="pstart"))
    await call.answer()


@dp.callback_query(F.data.startswith("pstart_day:"), StateFilter(AdminForm.picking_period))
async def period_start_calendar_pick_day(call: CallbackQuery, state: FSMContext):
    _, date_str = call.data.split(":", 1)
    await state.update_data(period_start=date_str)
    year_s, month_s, day_s = date_str.split("-")

    await call.message.edit_reply_markup(reply_markup=None)
    await call.message.answer(
        f"Начало периода: {day_s}.{month_s}.{year_s}.\n"
        "Теперь выберите дату ОКОНЧАНИЯ периода:",
        reply_markup=_build_calendar(int(year_s), int(month_s), prefix="pend"),
    )
    await call.answer()


@dp.callback_query(F.data.startswith("pstart_day:"))
async def period_start_calendar_pick_day_stale(call: CallbackQuery):
    await call.answer("Эта кнопка уже не активна. Начните заново через меню «📊 Отчёты» / «🛠 Модерация».", show_alert=True)


@dp.callback_query(F.data == "pend_ignore")
async def period_end_calendar_ignore(call: CallbackQuery):
    await call.answer()


@dp.callback_query(F.data.startswith("pend_nav:"), StateFilter(AdminForm.picking_period))
async def period_end_calendar_nav(call: CallbackQuery):
    _, ym = call.data.split(":", 1)
    year, month = map(int, ym.split("-"))
    await call.message.edit_reply_markup(reply_markup=_build_calendar(year, month, prefix="pend"))
    await call.answer()


@dp.callback_query(F.data.startswith("pend_day:"), StateFilter(AdminForm.picking_period))
async def period_end_calendar_pick_day(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    start_str = data.get("period_start")
    purpose = data.get("period_purpose")
    if not start_str or not purpose:
        await call.answer("Сначала выберите дату начала периода. Начните заново.", show_alert=True)
        await state.clear()
        return

    _, end_str = call.data.split(":", 1)
    start_year_s, start_month_s, start_day_s = start_str.split("-")
    end_year_s, end_month_s, end_day_s = end_str.split("-")
    start_date = date(int(start_year_s), int(start_month_s), int(start_day_s))
    end_date = date(int(end_year_s), int(end_month_s), int(end_day_s))

    if end_date < start_date:
        await call.answer("Дата окончания раньше даты начала — выберите другую дату.", show_alert=True)
        return

    await call.message.edit_reply_markup(reply_markup=None)
    await call.answer()
    await call.message.answer(
        f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}. "
        "Сверяю данные, секунду..."
    )
    await _run_period_matching(call.message, state, purpose, start_date, end_date, call.from_user.id)


@dp.callback_query(F.data.startswith("pend_day:"))
async def period_end_calendar_pick_day_stale(call: CallbackQuery):
    await call.answer("Эта кнопка уже не активна. Начните заново через меню «📊 Отчёты» / «🛠 Модерация».", show_alert=True)


@dp.message(StateFilter(AdminForm.picking_period))
async def wrong_content_for_period_picking(message: Message):
    await message.answer("Пожалуйста, выберите дату кнопками календаря в сообщении выше.")


async def _run_period_matching(
    message: Message,
    state: FSMContext,
    purpose: str,
    start_date: date,
    end_date: date,
    user_id: int,
):
    """Общая логика для трёх пунктов, использующих выбор периода двумя
    календарями: «📊 Отчёты -> Отчёт по продажам», «📊 Отчёты -> Отчёт по
    пользователям» и «🛠 Модерация -> Автоматическая модерация»."""
    filename_suffix = f"{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}"

    if purpose == "user_report":
        try:
            buf, rows_count = _build_user_report_workbook(start_date, end_date)
        except Exception:
            logger.exception("Не удалось сформировать отчёт по пользователям")
            await message.answer("Не получилось сформировать отчёт, попробуйте позже.")
            await state.clear()
            return

        await state.clear()
        filename = f"otchet_polzovateli_{filename_suffix}.xlsx"
        await message.answer_document(
            BufferedInputFile(buf.read(), filename=filename),
            caption=f"Строк в сводке: {rows_count}.",
            reply_markup=_menu_for(user_id),
        )
        return

    try:
        buf, matched_count, unmatched_count, matched_rows = _build_sales_match_workbook(start_date, end_date)
    except Exception:
        logger.exception("Не удалось выполнить сверку продаж")
        await message.answer("Не получилось выполнить сверку, попробуйте позже.")
        await state.clear()
        return

    if purpose == "auto_mod":
        filename = f"avto_moderatsiya_{filename_suffix}.xlsx"
        await message.answer_document(
            BufferedInputFile(buf.read(), filename=filename),
            caption=(
                f"Совпадений: {matched_count}. Без совпадений: {unmatched_count}. "
                f"Уникальных чеков с совпадением: {len(matched_rows)}."
            ),
        )

        if not matched_rows:
            await state.clear()
            await message.answer("Менять нечего — совпадений не найдено.", reply_markup=_menu_for(user_id))
            return

        await state.set_state(AdminForm.auto_mod_confirm)
        await state.update_data(matched_rows=matched_rows)
        kb = InlineKeyboardBuilder()
        kb.row(
            InlineKeyboardButton(text="✅ Да", callback_data="auto_mod_apply_yes"),
            InlineKeyboardButton(text="❌ Нет", callback_data="auto_mod_apply_no"),
        )
        await message.answer(
            f"Изменить статус чеков на «принят» у {len(matched_rows)} шт. "
            "(совпавших по ИНН/№ УПД/дате продажи с акционным товаром)? "
            "Пользователям сообщение отправлено НЕ будет — купон в этом "
            "сценарии не назначается, меняется только статус в таблице.",
            reply_markup=kb.as_markup(),
        )
        return

    # purpose == "sales_match"
    await state.clear()
    filename = f"sverka_{filename_suffix}.xlsx"
    await message.answer_document(
        BufferedInputFile(buf.read(), filename=filename),
        caption=f"Совпадений: {matched_count}. Без совпадений: {unmatched_count}.",
        reply_markup=_menu_for(user_id),
    )


@dp.callback_query(F.data == "auto_mod_apply_yes", IsModerator(), StateFilter(AdminForm.auto_mod_confirm))
async def admin_moderation_auto_apply_yes(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    rows = data.get("matched_rows", [])
    await state.clear()
    await call.answer()

    updated = 0
    for row in rows:
        try:
            google_sheets.update_receipt_status(
                row,
                google_sheets.STATUS_ACCEPTED,
                comment="Подтверждено автоматической сверкой с отчётом по продажам",
            )
            updated += 1
        except Exception:
            logger.exception("Не удалось обновить статус чека %s при автомодерации", row)

    await call.message.answer(f"Готово, статус изменён у {updated} из {len(rows)} чеков.")


@dp.callback_query(F.data == "auto_mod_apply_no", IsModerator(), StateFilter(AdminForm.auto_mod_confirm))
async def admin_moderation_auto_apply_no(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.answer()
    await call.message.answer("Хорошо, статусы не менял.")


@dp.callback_query(F.data.in_({"auto_mod_apply_yes", "auto_mod_apply_no"}))
async def admin_moderation_auto_apply_stale(call: CallbackQuery):
    await call.answer("Эта кнопка уже не активна.", show_alert=True)


async def main():
    # На Amvera (и вообще при передеплое на любом хостинге) старый процесс
    # бота может ещё секунду-другую держать соединение с Telegram после
    # того, как запустился новый. Без небольшой паузы это иногда даёт
    # ошибку "Terminated by other getUpdates request".
    startup_delay = float(os.getenv("STARTUP_DELAY_SECONDS", "3"))
    if startup_delay > 0:
        await asyncio.sleep(startup_delay)

    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
